# Copyright (c) 2020-2021 Molly White
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from openpyxl import load_workbook


def get_excel_data_for_date_range(filename, date_range, sheetname=None):
    """Gets the rows for the selected date range, represented as a dict. This makes some
    assumptions about the format of the Excel sheet, including that the first row is
    the header row and that the first column contains the dates."""
    wb = load_workbook(filename=filename, data_only=True)

    # If sheet name isn't passed in, assume we're looking at the first sheet
    if not sheetname:
        sheets = wb.sheetnames
        sheetname = sheets[0]
    sheet = wb[sheetname]

    headings = []
    result = {d: None for d in date_range}

    for ind, row in enumerate(sheet.rows):
        if ind == 0:
            headings = [x.value for x in row]
        else:
            dt = row[0].value
            if dt and dt.date() in date_range:
                result[dt.date()] = {
                    x: row[h_ind].value for h_ind, x in enumerate(headings)
                }

    return result


def safe_lookup(excel_dict, key, default=None):
    if type(excel_dict) is dict and key in excel_dict:
        try:
            val = int(excel_dict[key])
            return val
        except ValueError:
            return excel_dict[key]
    return default
